from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only = True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져오기
for row in ws.values:
    for cell in row:
        print(cell)