from re import A
from openpyxl import load_workbook
wb = load_workbook("question.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2,min_col=4, max_col=4):
    for cell in row:
        cell.value = 10

ws["H1"].value = "총점"
ws["I1"].value = "성적"

i = 2
j = 2
sum_val = 0 


for row in ws.iter_rows(min_row=2,min_col=2, max_col=9):
    row[6].value = "=SUM(B{0}:G{1})".format(i,j)
    i += 1
    j += 1
    sum_val = 0

    for a in range(0,6):
        sum_val += row[a].value
    print(sum_val)

    
    if row[0].value < 5:
        row[7].value = "F"
        continue
        
    if sum_val >= 90:
        row[7].value = "A"
        
    elif sum_val >= 80:
        row[7].value = "B"

    elif sum_val >= 70:
        row[7].value = "C"
    else:
        row[7].value = "D"

wb.save("scores.xlsx")

# H열에 총점(SUM), I열에 성적 정보 추가 (90A 80B, 70C, other D)
# 출석 5 미만인 학생은 F
# 최종 파일 : scores.xlsx